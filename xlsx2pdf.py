import argparse
import datetime
import functools
import os
import re
import textwrap

import openpyxl
import unicodedata
import win32com.client as win32
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.pagebreak import Break

height = 35
m_height = 28
s_height = 22

# 定义位置样式
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

# 定义字体样式
bold_font = Font(bold=True, size=14, name='等线')
s_bold_font = Font(bold=True, size=10, name='微软雅黑')
simp_font = Font(bold=False, size=10, name='微软雅黑')

# 定义背景样式
gree_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
blue_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')

# 定义单元格样式
cell_border = Border(
	left=Side(style='thin', color='000000'),
	right=Side(style='thin', color='000000'),
	top=Side(style='thin', color='000000'),
	bottom=Side(style='double', color='000000')
)

edge = Border(
	left=Side(style='thick', color='000000'),
	right=Side(style='thick', color='000000'),
	top=Side(style='thick', color='000000'),
	bottom=Side(style='thick', color='000000')
)

inside = Border(
	left=Side(style='dotted', color='A9A9A9'),
	right=Side(style='dotted', color='A9A9A9'),
	top=Side(style='dotted', color='A9A9A9'),
	bottom=Side(style='dotted', color='A9A9A9')
)


def regular_border(ws, start_row, end_row, start_col=1, end_col=6):
	# 遍历指定范围内的每个单元格，并应用样式
	for row in ws.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
		for cell in row:
			# 应用单元格样式
			if cell.row == start_row and cell.column != start_col and cell.column != end_col:
				cell.border += cell_border
			if cell.row == end_row and cell.column != start_col and cell.column != end_col:
				cell.border += cell_border
			elif cell.column != start_col and cell.column != end_col:
				cell.border += inside

		for cell in row:
			if cell.column == start_col:
				cell.border += Border(
					left=edge.left, 
					right=cell_border.right,
					top=cell_border.top,
					bottom=cell_border.bottom
					)
			if cell.column == end_col:
				cell.border += Border(
					left=cell_border.left, 
					right=edge.right,
					top=cell_border.top,
					bottom=cell_border.bottom
					)


def regular_border_inside(ws, start_row, end_row, start_col=1, end_col=6):
	for row in range(start_row + 1, end_row):
		for col in range(start_col + 1, end_col):
			cell = ws.cell(row=row, column=col)
			cell.border += inside

	for row in range(start_row, end_row + 1):
		for col in range(start_col, end_col + 1):
			cell = ws.cell(row=row, column=col)
			if row == start_row and col == start_col:
				cell.border += Border(
					top=edge.top,
					bottom=inside.bottom,
					left=edge.left,
					right=inside.right
				)
			if row == start_row and col == end_col:
				cell.border += Border(
					top=edge.top,
					bottom=inside.bottom,
					left=inside.left,
					right=edge.right
				)
			if row == end_row and col == start_col:
				cell.border += Border(
					top=inside.top,
					bottom=edge.bottom,
					left=edge.left,
					right=inside.right
				)
			if row == end_row and col == end_col:
				cell.border += Border(
					top=inside.top,
					bottom=edge.bottom,
					left=inside.left,
					right=edge.right
				)
			if row == start_row and col != start_col:
				cell.border += Border(
					top=edge.top,
					bottom=inside.bottom,
					left=inside.left,
					right=inside.right
				)
			if row == end_row and col != end_col:
				cell.border += Border(
					top=inside.top,
					bottom=edge.bottom,
					left=inside.left,
					right=inside.right
				)
			if col == start_col and row != start_row:
				cell.border += Border(
					top=inside.top,
					bottom=inside.bottom,
					left=edge.left,
					right=inside.right
				)
			if col == end_col and row != end_row:
				cell.border += Border(
					top=inside.top,
					bottom=inside.bottom,
					left=inside.left,
					right=edge.right
				)


# 遍历所有行并查找目标内容
def get_row(ws, start_row, target):
	for row in ws.iter_rows(min_row=start_row, max_col=ws.max_column, max_row=ws.max_row):
		for cell in row:
			if cell.value == target:
				return cell.row
	return -1


def cal_height(ws, row):
	max_line_cnt = 0
	for cell in ws[row]:
		line_cnt = 0
		if cell.value is not None and type(cell.value) == str:
			paragraphs = cell.value.split('\n')
			for p in paragraphs:  # 如果段落不为空，则考虑其中的所有行
				if len(p.strip()) > 0:  # 将段落按照自动换行进行拆分成多个行
					lines = textwrap.wrap(p, width=60)
					line_cnt += len(lines)
			max_line_cnt = max(max_line_cnt, line_cnt)

	if max_line_cnt * s_height < m_height:
		return m_height
	return max_line_cnt * s_height + 14


def deal_excel(excel_path, day):
	data_dict = {"info": [], "本周工作总结": [], "下周工作计划": [], "备注事项": []}

	wb = openpyxl.load_workbook(excel_path)
	ws = wb.active

	info_row = get_row(ws, 0, "姓名")
	if info_row == -1:
		print("error: cannot find 姓名 in sheet {}".format(excel_path))
		return None

	summer_row = get_row(ws, info_row, "本周工作总结")
	if summer_row == -1:
		print("error: cannot find 本周工作总结 in sheet {}".format(excel_path))
		return None

	plan_row = get_row(ws, summer_row, "下周工作计划")
	if plan_row == -1:
		print("error: cannot find 下周工作计划 in sheet {}".format(excel_path))
		return None

	backup_row = get_row(ws, plan_row, "备注事项")
	if backup_row == -1:
		print("error: cannot find 备注事项 in sheet {}".format(excel_path))
		return None

	for row in ws.iter_rows(min_row=info_row, max_col=ws.max_column, max_row=summer_row - 1):
		for cell in row:
			if cell.value is not None:
				data_dict["info"].append(cell.value)

		# if len(data_dict["info"]) == 6:
		# 	continue
		if len(data_dict["info"]) == 5:
			data_dict["info"].append(day)
		elif len(data_dict["info"]) < 5:
			print("error: miss data in personal info: ", data_dict["info"])
			return None

	for row in ws.iter_rows(min_row=summer_row + 1, max_col=ws.max_column, max_row=plan_row - 1):
		tmp = []
		for cell in row:
			if cell.value is not None:
				tmp.append(cell.value)
		data_dict["本周工作总结"].append(tmp)

	for row in ws.iter_rows(min_row=plan_row + 1, max_col=ws.max_column, max_row=backup_row - 1):
		tmp = []
		for cell in row:
			if cell.value is not None:
				tmp.append(cell.value)
		data_dict["下周工作计划"].append(tmp)

	for row in ws.iter_rows(min_row=backup_row + 1, max_col=ws.max_column, max_row=ws.max_row):
		tmp = []
		for cell in row:
			if cell.value is not None:
				tmp.append(cell.value)
		data_dict["备注事项"].append(tmp)

	wb.close()

	# print(data_dict)
	return data_dict


def re_cn_punctuation(text):
	if type(text) != str:
		return text

	# 中文全角符号 Unicode 范围：\uFF00-\uFF5F
	chinese_punc = [chr(i) for i in range(0xFF01, 0xFF5F + 1)]

	# 英文半角符号对应 Unicode 范围：\u0021-\u007E
	english_punc = "!?.,\"#$%&'()*+,-/:;<=>@[\\]^_`{|}~"

	# 将中文全角符号替换为英文半角符号
	for c, e in zip(chinese_punc, english_punc):
		text = unicodedata.normalize("NFKC", text).replace(c, e)

	text = re.sub(" ", "", text)

	return text


def remove_file(file_path):
	if os.path.exists(file_path):
		if os.path.isfile(file_path):
			os.remove(file_path)
			return True
		else:
			print("{} is a dir path, cannot remove it".format(file_path))
			return False
	return True


def sort_string_array(strs, order):
	# 自定义比较函数
	def compare(s1, s2):
		a , b = 0, 0
		for i, char in enumerate(order):
			if char in s1:
				a = i
			if char in s2:
				b = i
    
		return a - b

	# 使用自定义比较函数排序
	strs.sort(key=functools.cmp_to_key(compare))

	return strs


def merge_excel(excel_dir_path, out_excel_path, order, week_range, day):
	if remove_file(out_excel_path) is False:
		return

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "mergeSummer"

	for col in range(1, 7):
		column_letter = ws.cell(row=1, column=col).column_letter
		if col == 1:
			ws.column_dimensions[column_letter].width = 12
		else:
			ws.column_dimensions[column_letter].width = 30
	row = 1

	ws.cell(row, 1, "华东大区工作周报汇总")
	ws.cell(row, 4, week_range)
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
	ws.row_dimensions[row].height = height
	for cell in ws[row]:
		cell.alignment = align_center
		cell.font = bold_font
		cell.fill = blue_fill

	regular_border(ws, row, row)

	row = add_black(ws, row, 2)

	name_list = os.listdir(excel_dir_path)
	name_list = sort_string_array(name_list, order)

	for name in name_list:
		excel_path = os.path.join(excel_dir_path, name)
		if not name.endswith(".xlsx") and os.path.isfile(excel_path):
			continue

		print("start to deal excel {}".format(name))
		data_dict = deal_excel(excel_path, day)
		if data_dict is None:
			print("failed to deal excel {}".format(name))
			continue

		for idx, val in enumerate(data_dict["info"]):
			ws.cell(row, idx + 1, re_cn_punctuation(val))
			ws.cell(row, idx + 1).alignment = align_center
			ws.cell(row, idx + 1).font = simp_font
			if idx == 1:
				ws.cell(row, idx + 1).font = s_bold_font
		for cell in ws[row]:
			cell.fill = blue_fill
		ws.row_dimensions[row].height = height

		regular_border(ws, row, row)

		row = write_subtitle(ws, row, "本周工作总结")
		start_row = row
		row = write_work(ws, row, data_dict["本周工作总结"], 1)

		row = write_subtitle(ws, row, "下周工作计划")
		row = write_work(ws, row, data_dict["下周工作计划"], 2)

		row = write_subtitle(ws, row, "备注事项")
		row = write_backup(ws, row, data_dict["备注事项"])

		regular_border_inside(ws, start_row, row)
		row = add_black(ws, row, 1)
		page_break = Break(id=row - 1)
		ws.row_breaks.append(page_break)

	for row in range(1, ws.max_row, 1):
		if ws.row_dimensions[row].height is None:
			ws.row_dimensions[row].height = cal_height(ws, row)

	wb.save(out_excel_path)
	wb.close()


def add_black(ws, row, line_num):
	ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + line_num, end_column=6)
	return row + line_num + 1


def write_subtitle(ws, row, subtitle):
	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
	ws.cell(row, 1, subtitle)

	ws.row_dimensions[row].height = height
	for cell in ws[row]:
		cell.alignment = align_center
		cell.font = s_bold_font
		cell.fill = gree_fill
	return row


def write_work(ws, row, data_list, order):
	for item in data_list:
		if len(item) == 0:
			continue

		row += 1
		if order == 1:
			ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
			ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
		elif order == 2:
			ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
			ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)

		ws.cell(row, 1, item[0])
		regular_cell(ws, row, 1, align_center, simp_font)

		if len(item) > 1:
			ws.cell(row, 2, item[1])
			regular_cell(ws, row, 2, align_left, simp_font)

		if len(item) > 2:
			if order == 1:
				ws.cell(row, 4, item[2])
				regular_cell(ws, row, 4, align_center, simp_font)
			elif order == 2:
				ws.cell(row, 4, item[2])
				regular_cell(ws, row, 4, align_left, simp_font)

		if len(item) > 3:
			if order == 1:
				ws.cell(row, 5, item[3])
				regular_cell(ws, row, 5, align_left, simp_font)
			elif order == 2:
				ws.cell(row, 6, item[3])
				regular_cell(ws, row, 6, align_center, simp_font)

		if item[0] == "序号":
			ws.row_dimensions[row].height = height
			for cell in ws[row]:
				cell.alignment = align_center
				cell.font = s_bold_font

	return row


def regular_cell(ws, row, col, align, font):
	ws.cell(row, col).alignment = align
	ws.cell(row, col).font = font


def write_backup(ws, row, data_list):
	for item in data_list:
		if len(item) == 0:
			continue
		row += 1
		ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
		ws.cell(row, 1, item[0])
		regular_cell(ws, row, 1, align_center, simp_font)
		if len(item) > 1:
			ws.cell(row, 2, item[1])
			regular_cell(ws, row, 2, align_left, simp_font)

		if item[0] == "序号":
			ws.row_dimensions[row].height = 35
			for cell in ws[row]:
				cell.alignment = align_center
				cell.font = s_bold_font

	return row


def xlsx2xlsm(in_xlsx_path, out_pdf_path, vba_path):
	xlsm_path = re.sub("xlsx", "xlsm", in_xlsx_path)
	if remove_file(xlsm_path) is False or remove_file(out_pdf_path) is False:
		return
	# 打开Excel应用程序
	excel_app = win32.gencache.EnsureDispatch('Excel.Application')
	excel_app.Visible = False

	wb = excel_app.Workbooks.Open(os.path.abspath(in_xlsx_path))
	# 保存为xlsm文件格
	wb.SaveAs(os.path.abspath(xlsm_path), FileFormat=52)
	# remove_file(in_xlsx_path)
	wb.Close()

	with open("D:/merge-xlsx/merge.vba", 'r', encoding='utf-8') as f:
		contents = f.read()

	with open(vba_path, 'w', encoding='utf-8') as f:
		re_sub = re.sub("520GIFTForL2023", out_pdf_path, contents, count=1)
		f.write(re_sub)


def get_time():
	# 获取当前时间和日期（北京时间）
	now = datetime.datetime.utcnow() + datetime.timedelta(hours=8)

	# 计算本周开始日期和结束日期
	start_of_week = now - datetime.timedelta(days=now.weekday())
	end_of_week = start_of_week + datetime.timedelta(days=6)

	# 格式化日期字符串
	date_format = '%Y%m%d'
	start_date_str = start_of_week.strftime(date_format)
	end_date_str = end_of_week.strftime(date_format)
	fri_date_str = start_of_week + datetime.timedelta(days=4)

	# 将日期字符串合并为日期范围字符串
	week_range = f'{start_date_str}-{end_date_str}'
	fri_day = fri_date_str.strftime('%Y年%m月%d日')
	
	return week_range, fri_day


def main():
	week_range, fri_day = get_time()
	file_name = "工作周报-部门汇总-{}".format(week_range)

	parser = argparse.ArgumentParser(description="A tool merging excels to xlsx and pdf.")

	# 添加命令行参数
	parser.add_argument("-f", "--folder", default="D:/work/week-report", help="folder path for excels")
	parser.add_argument("-o", "--order", default="林鲁单冀坤朱茗马涂", help="excel deal order")
	parser.add_argument("-x", "--xlsx", default="D:/work/week-report-output/{}.xlsx".format(file_name), help="output excel path")
	parser.add_argument("-p", "--pdf", default="D:/work/week-report-output/{}.pdf".format(file_name), help="output pdf path")
	parser.add_argument("-v", "--vba", default="D:/work/week-report-output/{}.vba".format(file_name), help="output vba path")
	parser.add_argument("-w", "--week", default=week_range, help="week range")
	parser.add_argument("-d", "--day", default=fri_day, help="personal day")

	# 解析命令行参数
	args = parser.parse_args()

	merge_excel(args.folder, args.xlsx, args.order, args.week, args.day)
	xlsx2xlsm(args.xlsx, args.pdf, args.vba)


if __name__ == "__main__":
	main()
